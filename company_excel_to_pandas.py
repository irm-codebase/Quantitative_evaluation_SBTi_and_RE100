from os import walk
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from math import isclose


PATH = "/home/ivan/Documents/GitHub/G500_database/data/"
COMPANIES = "/home/ivan/Documents/GitHub/G500_database/data/companies/collected/"


def clear_databases():
    """
    Hard cleaning of the entire database.
    :return:
    """
    # RE100 is not cleaned because it has pre-existing values that cannot be obtained in automatized ways

    # Clean all values in the emissions database
    df = pd.read_csv(PATH + 'G500_emissions.csv', index_col=0, header=[0])
    df.loc[:] = np.nan
    df.to_csv(PATH + 'G500_emissions.csv')

    # Clean all values in the energy database
    df = pd.read_csv(PATH + 'G500_energy.csv', index_col=0, header=[0])
    df.loc[:] = np.nan
    df.to_csv(PATH + 'G500_energy.csv')

    # Clean 'completed' from central database
    df = pd.read_csv(PATH + 'G500_central.csv', index_col=0, header=[0])
    df['completed'] = np.nan
    df.to_csv(PATH + 'G500_central.csv')

    # Clean all values in S2MB database
    df = pd.read_csv(PATH+'G500_S2MB.csv', index_col=0, header=[0])
    df.loc[:] = np.nan
    df.to_csv(PATH+'G500_S2MB.csv')

    # Clean absolute targets from sbti database
    df = pd.read_csv(PATH + 'G500_sbti.csv', index_col=0, header=[0])
    df.loc[:, 'start year t1':] = np.nan
    df.to_csv(PATH + 'G500_sbti.csv')

    # Clean all RE100 target data
    df = pd.read_csv(PATH + 'G500_re100.csv', index_col=0, header=[0])
    df.loc[:, 'joining year':] = np.nan
    df.to_csv(PATH + 'G500_re100.csv')


def build_re100_database(name, worksheet=None):
    """
    Fetch data from each company's file
    :param name: name of the company (used as index)
    :param worksheet: excel worksheet to get data from
    :return:
    """
    re100_df = pd.read_csv(PATH + 'G500_re100.csv', index_col=0, header=[0])
    # Get report database and remove false 'spaces' just in case
    report_df = pd.read_excel(PATH + "other/RE100_report_2020.ods", index_col=0, header=[0])
    report_df = report_df.replace(' ', np.nan)

    df_membership = re100_df.loc[name, 'member']

    if df_membership:
        if name in report_df.index:
            # Target info
            target_info = report_df.loc[name]
            columns = ['joining year', 'final year', 'interim year 1', 'interim target 1', 'interim year 2',
                       'interim target 2']
            for c in columns:
                re100_df.loc[name, c] = target_info[c]
            re100_df.loc[name, 'final target'] = 1
            # Identify the earliest info in the database, if available.
            # Otherwise, pre-set baseline year only. These will have to be filled with other data.
            for year in range(2014, 2020):
                if not np.isnan(target_info[year]):
                    baseline_year = year
                    break
            else:
                if target_info['joining year'] == 2014:
                    baseline_year = 2014
                else:
                    baseline_year = target_info['joining year'] - 1
            re100_df.loc[name, 'baseline'] = target_info[baseline_year]
            re100_df.loc[name, 'baseline year'] = baseline_year
        else:
            # Fills target data but not baseline data
            ws_membership = worksheet['B12'].value
            if ws_membership == 'yes':
                for i in range(3, 5):
                    target_type = worksheet['D'+str(i)].value
                    if target_type == "Final":
                        re100_df.loc[name, 'joining year'] = int(worksheet['E' + str(i)].value)
                        re100_df.loc[name, 'final year'] = int(worksheet['F'+str(i)].value)
                        re100_df.loc[name, 'final target'] = float(worksheet['G' + str(i)].value)
                    elif target_type == "Interim 1":
                        re100_df.loc[name, 'interim year'] = int(worksheet['F'+str(i)].value)
                        re100_df.loc[name, 'interim target'] = float(worksheet['G' + str(i)].value)
                joining_yr = re100_df.loc[name, 'joining year']
                if joining_yr == 2014:
                    re100_df.loc[name, 'baseline year'] = 2014
                elif joining_yr == 2015:
                    re100_df.loc[name, 'baseline year'] = 2015
                else:
                    re100_df.loc[name, 'baseline year'] = joining_yr - 1
            elif pd.isna(df_membership) and ws_membership is None:
                return
            else:
                raise ValueError(name, "RE100 membership mismatch. Excel ", ws_membership, "Database", df_membership)
    re100_df.to_csv(PATH + 'G500_re100.csv')


def fill_missing_re100():
    missing = ["Seven & I Holdings", "M&G"]

    for name in missing:
        build_re100_database(name)


def check_re100_database():
    """
    Verify that RE100 targets are in order.
    :return:
    """
    re100_df = pd.read_csv(PATH + 'G500_re100.csv', index_col=0, header=[0])
    for index in re100_df.index:
        if re100_df.loc[index, 'member'] == 'yes':
            start = re100_df.loc[index, 'joining year']
            interim1 = re100_df.loc[index, 'interim year 1']
            interim2 = re100_df.loc[index, 'interim year 2']
            final = re100_df.loc[index, 'final year']
            f_target = re100_df.loc[index, 'final target']
            i_target1 = re100_df.loc[index, 'interim target 1']
            i_target2 = re100_df.loc[index, 'interim target 2']
            if not (2014 <= start <= final <= 2050):
                raise ValueError(index, "start/final year error. Start year", start, "Final year", final)
            if not (f_target == 1):
                raise ValueError(index, "Final target is not 100%.")
            if not pd.isnull(interim1):
                if not (2015 < interim1 < final):
                    raise ValueError(index, "Interim 1 year error", interim1)
                if not (0 < i_target1 < 1):
                    raise ValueError(i_target1, "Interim 1 target error", i_target1)
            if not pd.isnull(interim2):
                if not (interim1 < interim2 < final):
                    raise ValueError(index, "Interim 1 year error", interim2)
                if not (i_target1 < i_target2 < 1):
                    raise ValueError(i_target2, "Interim 1 target error", i_target2)


def build_emissions_database(name, worksheet):
    """
    Fetch data from each company's file
    :param name: name of the company (used as index)
    :param worksheet: excel worksheet to get data from
    :return:
    """
    n_years = 5

    emissions_df = pd.read_csv(PATH + 'G500_emissions.csv', index_col=0, header=[0])
    scopes_rows = {'S1': 19, 'S2 LB': 20, 'S2 MB': 21, 'S3': 23, 'S3 C1': 24, 'S3 C2': 25, 'S3 C3': 26, 'S3 C4': 27,
                   'S3 C5': 28, 'S3 C6': 29, 'S3 C7': 30, 'S3 C8': 31, 'S3 C9': 32, 'S3 C10': 33, 'S3 C11': 34,
                   'S3 C12': 35, 'S3 C13': 36, 'S3 C14': 37, 'S3 C15': 38, 'S3 other (upstream)': 39,
                   'S3 other (downstream)': 40, 'S3 sector specific': 41}

    for i in range(n_years):
        col = chr(ord('E')+i)
        for scope in scopes_rows:
            row = str(scopes_rows[scope])
            val = worksheet[col+row].value
            if val:
                emissions_df.loc[name, scope + ' ' + str(2015 + i)] = val

    emissions_df.to_csv(PATH + 'G500_emissions.csv')


def check_emissions_database():
    """
    Check if emissions database makes sense.
    Tests:
    S3 total == sum of categories (if disclosed)
    :return:
    """
    emissions_df = pd.read_csv(PATH + 'G500_emissions.csv', index_col=0, header=[0])

    # S3 total == sum of categories (if disclosed)
    for i in range(5):
        # Scope 3 total dataframe
        year = " " + str(2015+i)
        s3_total_df = emissions_df['S3' + year]
        s3_total_df.fillna(0)

        # Scope 3 categories dataframe
        s3_cat_sum_df = emissions_df.loc[:, "S3 C1"+year:"S3 sector specific"+year].sum(axis=1)

        for index in emissions_df.index:
            if not isclose(s3_total_df[index], s3_cat_sum_df[index]):
                if s3_cat_sum_df[index]:
                    print(index, "Scope 3 mismatch error. Total", s3_total_df[index], "Sum", s3_cat_sum_df[index])


def build_energy_database(name, worksheet):
    """
    Fetch data from each company's file
    :param name: name of the company (used as index)
    :param worksheet: excel worksheet to get data from
    :return:
    """
    # Don't bother gathering data for utility companies, because they do not understand the concept of Scope 2 >:(
    utilities = ['Enel', 'Electricité de France', 'Engie', 'Iberdrola', 'Naturgy Energy Group', 'Veolia Environnement']
    if name in utilities:
        return

    n_years = 5

    energy_df = pd.read_csv(PATH + 'G500_energy.csv', index_col=0, header=[0])
    data = energy_df.columns.to_list()
    data = data[0:21]
    data = [c.replace(' 2015', '') for c in data]

    for i in range(n_years):
        col = chr(ord('E')+i)
        for j in range(len(data)):
            if j < 13:
                row = str(12 + j)
            elif j < 17:
                row = str(15 + j)
            elif i > 1:
                row = str(17 + j)
            else:
                break

            val = worksheet[col+row].value
            if val is not None:
                energy_df.loc[name, data[j] + ' ' + str(2015 + i)] = float(val)
            else:
                raise ValueError(name, 'empty energy section in cell', col, row, 'value', val)

    energy_df.to_csv(PATH + 'G500_energy.csv')


def check_energy_database():
    """
    Validate energy data. Same tests as the Excel. Just in case.
    :return:
    """
    energy_df = pd.read_csv(PATH + 'G500_energy.csv', index_col=0, header=[0])
    energy_df = energy_df.fillna(0)

    for i in range(2015, 2020):
        year = ' ' + str(i)

        # consumption series
        sum_fuel = energy_df['cr fuel' + year] + energy_df['cnr fuel' + year]
        sum_p_el = energy_df['cr purchased electricity'+year] + energy_df['cnr purchased electricity'+year]
        sum_p_hsc = energy_df['cr purchased hsc'+year] + energy_df['cnr purchased hsc'+year]
        sum_renewable = (energy_df['cr fuel'+year] + energy_df['cr purchased electricity'+year] +
                         energy_df['cr purchased hsc'+year] + energy_df['cr self-gen non-fuel'+year])
        sum_nonrenewable = (energy_df['cnr fuel'+year] + energy_df['cnr purchased electricity'+year] +
                            energy_df['cnr purchased hsc'+year])
        sum_total = energy_df['cr energy'+year] + energy_df['cnr energy'+year]

        # generation series
        gross_el = energy_df['gt gross electricity'+year]
        self_cons_el = energy_df['gt self-cons electricity'+year]
        gross_ren_el = energy_df['gr gross electricity'+year]
        self_cons_ren_el = energy_df['gr self-cons electricity'+year]

        gross_hsc = energy_df['gt gross hsc' + year]
        self_cons_hsc = energy_df['gt self-cons hsc' + year]
        gross_ren_hsc = energy_df['gr gross hsc' + year]
        self_cons_ren_hsc = energy_df['gr self-cons hsc' + year]

        for index in energy_df.index:
            # consumption checks
            if not isclose(energy_df.loc[index, 'ct fuel'+year], sum_fuel[index]):
                print(index, 'consumed fuel error', year)
            if not isclose(energy_df.loc[index, 'ct purchased electricity'+year], sum_p_el[index]):
                print(index, 'consumed purchased electricity error', year)
            if not isclose(energy_df.loc[index, 'ct purchased hsc'+year], sum_p_hsc[index]):
                print(index, 'consumed purchased hsc error', year)
            if not isclose(energy_df.loc[index, 'cr energy'+year], sum_renewable[index]):
                print(index, 'consumed renewable energy error', year)
            if not isclose(energy_df.loc[index, 'cnr energy'+year], sum_nonrenewable[index]):
                print(index, 'consumed non-renewable energy error', year)
            if not isclose(energy_df.loc[index, 'ct energy'+year], sum_total[index]):
                print(index, 'consumed total energy error', year)
            # generation checks
            if gross_el[index] < self_cons_el[index] or gross_hsc[index] < self_cons_hsc[index]:
                print(index, 'Gross generation smaller than self-consumption', year)
            if gross_el[index] < gross_ren_el[index] or gross_hsc[index] < gross_ren_hsc[index]:
                print(index, 'Gross generation smaller than gross renewable', year)
            if self_cons_el[index] < self_cons_ren_el[index] or self_cons_hsc[index] < self_cons_ren_hsc[index]:
                print(index, 'Self-consumption smaller than renewable self-consumption', year)
            if gross_ren_el[index] < self_cons_ren_el[index] or gross_ren_hsc[index] < self_cons_ren_hsc[index]:
                print(index, 'Gross renewable smaller than renewable self-consumption', year)
            # consumption vs generation checks
            if not isclose(energy_df.loc[index, 'cr self-gen non-fuel'+year], self_cons_ren_el[index]):
                if energy_df.loc[index, 'cr energy'+year] < self_cons_ren_el[index] + self_cons_ren_hsc[index]:
                    print(index, 'consumed renewable generation bigger than total renewable consumption', year)


def build_check_central_database(name, worksheet):
    """
    Some tests on the central database. Specifically for sector values and completed flags.
    :param name:
    :param worksheet:
    :return:
    """
    central_df = pd.read_csv(PATH + 'G500_central.csv', index_col=0, header=[0])

    sector = worksheet['B5'].value
    industry = worksheet['B6'].value
    c_type = worksheet['B7'].value
    completed = worksheet['B8'].value

    if completed == 'yes':
        central_df.loc[name, 'completed'] = 'yes'
    else:
        raise ValueError(name, 'is not marked as completed')
    if sector != central_df.loc[name, 'sector']:
        raise ValueError(name, 'sector mismatch')
    if industry != central_df.loc[name, 'industry']:
        raise ValueError(name, 'industry mismatch')
    if c_type != central_df.loc[name, 'type']:
        raise ValueError(name, 'type mismatch')

    central_df.to_csv(PATH+'G500_central.csv')


def build_central_special_sectors():
    """
    Fill special sector data using Fortune data.
    Useful since the magazine did not separate sectors in a particularly useful way.
    :return:
    """
    central_df = pd.read_csv(PATH + 'G500_central.csv', index_col=0, header=[0])
    energy_sec = pd.read_csv(PATH + "mini databases/energy sectors.csv", index_col=0, header=[0])
    gics_sec = pd.read_csv(PATH + "mini databases/gics sectors.csv", index_col=0, header=[0])

    for index in central_df.index:
        industry = central_df.loc[index, 'industry']
        # Fill energy specific sector
        if industry in energy_sec.index.to_list():
            central_df.loc[index, "energy sector"] = energy_sec.loc[industry, "energy sector"]
        else:
            raise ValueError(index, industry, 'not found in energy sector database')
        # Fill GICS sector
        if industry in gics_sec.index.to_list():
            central_df.loc[index, "gics sector"] = gics_sec.loc[industry, "GICS"]
        else:
            raise ValueError(index, industry, 'not found in energy sector database')

    central_df.to_csv(PATH + 'G500_central.csv')


def build_s2mb_database(name, worksheet):
    """
    Fetch data from each company's file. Ignores utilities.
    :param name:
    :param worksheet:
    :return:
    """
    s2mb_df = pd.read_csv(PATH+'G500_S2MB.csv', index_col=0, header=[0])
    s2mb_df.loc[name].values[:] = 0

    utilities = ['Enel', 'Electricité de France', 'Engie', 'Iberdrola', 'Naturgy Energy Group']
    if name in utilities:
        return

    for i in range(5):
        year = ' ' + str(2015 + i)
        inst_col = chr(ord('A') + i*3)
        tech_col = chr(ord(inst_col) + 1)
        mwh_col = chr(ord(inst_col) + 2)

        row = 3
        while row < 47:
            instrument = worksheet[inst_col+str(row)].value
            if instrument:
                instrument = instrument.lower()
                mwh = float(worksheet[mwh_col+str(row)].value)
                try:
                    s2mb_df.loc[name, instrument+year] += mwh
                except KeyError:
                    raise KeyError(name, 'invalid instrument in year'+year)

                if i > 1:
                    techs = worksheet[tech_col+str(row)].value
                    techs = techs.lower()
                    techs = techs.split(';')
                    if techs[-1] == '':
                        techs.pop(-1)

                    for tech in techs:
                        try:
                            s2mb_df.loc[name, tech+year] += mwh/len(techs)
                        except KeyError:
                            raise KeyError(name, 'invalid technology in year'+year)
            else:
                break
            row += 1

        total = float(worksheet[chr(ord('Q')+i) + '3'].value)
        s2mb_df.loc[name, 'total'+year] = total

    s2mb_df.to_csv(PATH+'G500_S2MB.csv')


def check_s2mb_database():
    """
    Verifies S2MB data, and flags issues that need solving.
    :return:
    """
    s2mb_df = pd.read_csv(PATH + 'G500_S2MB.csv', index_col=0, header=[0])
    s2mb_df = s2mb_df.fillna(0)

    instruments = ['ppa direct line', 'ppa w/eac', 'ppa no eac', 'energy product w/eac',	'energy product no eac',
                   'unbundled eac', 'hsc agreement']

    techs = ['solar', 'wind', 'hydro', 'biomass', 'other tech',	'unspecified']

    for index in s2mb_df.index:
        for i in range(5):
            year = " " + str(2015+i)

            total = 0
            for instrument in instruments:
                total += s2mb_df.loc[index, instrument+year]
            if not isclose(total, s2mb_df.loc[index, 'total'+year]):
                raise ValueError(index, 'instrument sum error in'+year, 'Total instrument', total)

            if i > 1:
                total = 0
                for tech in techs:
                    total += s2mb_df.loc[index, tech + year]
                if not isclose(total, s2mb_df.loc[index, 'total' + year]):
                    raise ValueError(index, 'technology sum error in'+year, 'Total technology', total)


def build_sbti_database(name, worksheet):
    """
    Fetch data from each company's file
    :param name: company name for indexing
    :param worksheet: emissions worksheet
    :return:
    """
    sbti_df = pd.read_csv(PATH + 'G500_sbti.csv', index_col=0, header=[0])
    scope_dict = {"Scope 1": 'S1', "Scope 2 (location-based)": "S2 LB", "Scope 2 (market-based)": "S2 MB",
                  "Scope 1+2 (location-based)": "S1+2 LB", "Scope 1+2 (market-based)": 'S1+2 MB',
                  "Scope 3 (upstream)": "S3 up", "Scope 3 (downstream)": 'S3 down',
                  "Scope 3 (upstream & downstream)": 'S3', "Scope 3: Purchased goods and services": "S3 C1",
                  "Scope 3: Capital goods": "S3 C2",
                  "Scope 3: Fuel and energy-related activities (not included in Scopes 1 or 2)": 'S3 C3',
                  "Scope 3: Upstream transportation and distribution": "S3 C4",
                  "Scope 3: Waste generated in operations": "S3 C5", "Scope 3: Business travel": "S3 C6",
                  "Scope 3: Employee commuting": 'S3 C7', "Scope 3: Upstream leased assets": 'S3 C8',
                  "Scope 3: Downstream transportation and distribution": 'S3 C9',
                  "Scope 3: Processing of sold products": 'S3 C10', "Scope 3: Use of sold products": 'S3 C11',
                  "Scope 3: End-of-life treatment of sold products": 'S3 C12',
                  "Scope 3: Downstream leased assets": 'S3 C13', "Scope 3: Franchises": 'S3 C14',
                  "Scope 3: Investments": 'S3 C15'}

    # only fill in members with targets
    if not (sbti_df.loc[name, 'member'] == 'yes' and sbti_df.loc[name, 'status'] == 'Targets Set'):
        return

    n_targets = 7
    n_values = 7
    for i in range(n_targets):
        target = " t" + str(i+1)
        row = str(10 + i)

        value = worksheet['E' + row].value
        if value is None:
            empty_flag = True
        else:
            empty_flag = False
            sbti_df.loc[name, 'start year' + target] = int(value)

        for j in range(1, n_values):
            col = chr(ord('E') + j)
            value = worksheet[col+row].value
            if empty_flag:
                if value is not None:
                    print(name, "incomplete or false data found in target" + target)
            elif j == 1:
                sbti_df.loc[name, 'base year' + target] = int(value)
            elif j == 2:
                sbti_df.loc[name, 'target year' + target] = int(value)
            elif j == 3:
                sbti_df.loc[name, 'tco2e covered' + target] = float(value)
            elif j == 4:
                sbti_df.loc[name, '% scope covered' + target] = float(value)
            elif j == 5:
                sbti_df.loc[name, '% reduction' + target] = float(value)
            elif j == 6:
                sbti_df.loc[name, 'scope' + target] = scope_dict[value]

    sbti_df.to_csv(PATH+'G500_sbti.csv')


def check_sbti_database():
    """
    Run the following tests:
    Targets set:
            any targets? abs target complete?
            start year >= base year
            target year > start year
            0 < reduction <= 1
    Empty values -> Committed and non-members. Empty abs targets
    :return:
    """
    sbti_df = pd.read_csv(PATH + 'G500_sbti.csv', index_col=0, header=[0])
    central_df = pd.read_csv(PATH + 'G500_central.csv', index_col=0, header=[0])

    n_abs_targets = 7
    for index in sbti_df.index:
        company = sbti_df.loc[index]
        if company['member'] == 'yes':
            if company['status'] == 'Targets Set':
                if company['intensity target S1':'start year t1'].any():
                    for i in range(1, n_abs_targets+1):
                        abs_target = company['start year t'+str(i):'scope t'+str(i)]
                        if abs_target.any():
                            if not abs_target[0] >= abs_target[1]:
                                raise ValueError(index, i, "start year can not be earlier than base year")
                            if not abs_target[2] > abs_target[0]:
                                raise ValueError(index, i, "target year can not be earlier than start year")
                            if not (abs_target[3] > 0 and abs_target[4] > 0):
                                raise ValueError(index, i, "covered tco2e/% has to be larger than 0")
                            if not (0 < abs_target[5] <= 1):
                                raise ValueError(index, i, "reduction is not a percentage")
                            if type(abs_target[6]) is not str:
                                raise ValueError(index, i, "invalid scope")
                elif central_df.loc[index, 'completed'] == 'yes':
                    raise ValueError(index, 'no targets found')
                else:
                    print(index, 'has incomplete targets and information')
            elif ~np.isnan(company['qualification']) or company['date target update':].any():
                raise ValueError(index, 'found value instead of NaN')
        elif company['status':].any():
            raise ValueError(index, 'found value instead of NaN')


def build_company(file):
    """
    Extract data from the excel file of a specific company
    :param file:
    :return:
    """
    if '.xlsx' not in file:
        return
    wb = load_workbook(COMPANIES + file, read_only=True, data_only=True)
    emissions = wb['Emissions']
    energy = wb['Energy']
    s2mb = wb['S2 MB sourcing']
    name = file.replace('.xlsx', '')

    build_check_central_database(name, emissions)
    build_re100_database(name, energy)
    build_emissions_database(name, emissions)
    build_energy_database(name, energy)
    build_s2mb_database(name, s2mb)
    build_sbti_database(name, emissions)
    print(name, 'completed successfully')


def fill_cdp_sbti_links():
    """
    Adds missing links to CDP/SBTi websites to Excel files.
    RUNNING WILL DELETE EXCEL FORMULA RESULTS.
    YOU WILL HAVE TO MANUALLY OPEN AND SAVE EVERY FILE TO FIX.
    :return:
    """
    _, _, filenames = next(walk(COMPANIES))
    for file in filenames:
        if '.xlsx' not in file:
            return
        wb = load_workbook(COMPANIES + file)
        emissions = wb['Emissions']
        emissions['B25'].value = "https://www.cdp.net/en"
        emissions['B26'].value = "https://sciencebasedtargets.org/"
        wb.save(COMPANIES+file)


def build_databases(test_only=False):
    """
    Get all the data from the Excel files. Make some coffee while it runs.
    :param test_only: if True, the building step is skipped, only running validation.
    :return:
    """
    if not test_only:
        clear_databases()

        _, _, filenames = next(walk(COMPANIES))

        print(filenames)
        for file in filenames:
            build_company(file)

    fill_missing_re100()
    # build_central_special_sectors()
    print("\nALL DATABASES COMPLETED WITHOUT ERRORS\nTesting databases...\n")
    check_re100_database()
    check_emissions_database()
    check_energy_database()
    check_s2mb_database()
    check_sbti_database()


build_databases()