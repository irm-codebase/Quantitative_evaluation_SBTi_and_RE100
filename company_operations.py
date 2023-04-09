import pandas as pd
import shutil
import os.path
from openpyxl import load_workbook
import cdp_parser

DATABASE_PATH = '/home/ivan/Documents/GitHub/G500_database/data/'
COMPANY_PATH = DATABASE_PATH + 'companies/'
TEMPLATE = '/home/ivan/Documents/GitHub/G500_database/templates/New template.xlsx'
G500_PATH = DATABASE_PATH + 'G500_central.csv'


INFO_CELLS = {'Rank': 'B3', 'Country': 'B4', 'Sector': 'B5', 'Industry': 'B6', 'Type': 'B7'}
SBTI_CELLS = {'SBTi status': 'B12', 'SBTi qualification': 'B13'}
RE100_CELLS = {'member': 'B12'}


def prepare_company_file(name, safety_checks=True):
    """
    Creates a company file and pre-fills generic and initiative information.
    Also prints the initiative info to make fill-in easier.
    :param name:
    :param safety_checks:
    :return: path to the new file
    """
    # Create new file to begin fill-in process, with some safety checks
    g500_df = pd.read_csv(G500_PATH, index_col=0, header=[0])
    re100_df = pd.read_csv(DATABASE_PATH+'G500_re100.csv', index_col=0, header=[0])
    if name not in g500_df.index and safety_checks:
        print("This is not a valid company name.\nI did nothing!")
        return
    new_path = COMPANY_PATH+name+'.xlsx'
    if os.path.isfile(new_path) and safety_checks:
        print("A file for this company already exists pal.\nI did nothing!")
        return
    shutil.copy(TEMPLATE, new_path)

    wb = load_workbook(new_path)
    ws1 = wb[wb.sheetnames[0]]
    ws2 = wb[wb.sheetnames[1]]

    # Fill in generic info and save the file
    ws1['B2'] = name
    for cell_name in INFO_CELLS:
        ws1[INFO_CELLS[cell_name]] = g500_df.loc[name, cell_name.lower()]
    for cell_name in SBTI_CELLS:
        ws1[SBTI_CELLS[cell_name]] = g500_df.loc[name, cell_name.lower()]
    for cell_name in RE100_CELLS:
        ws2[RE100_CELLS[cell_name]] = re100_df.loc[name, cell_name.lower()]
    wb.save(new_path)

    # Tell me the targets so I can fill them in
    if g500_df.loc[name, 'sbti member'] == 'yes':
        print("SBTi target:")
        print(g500_df.loc[name, 'sbti status'])
        print(g500_df.loc[name, 'sbti target'])
    if re100_df.loc[name, 'member'] == 'yes':
        print('RE100 target')
        print(re100_df.loc[name, 'text'])

    return new_path


def parse_company(name, html_paths):
    excel = prepare_company_file(name)
    ordered_responses = cdp_parser.sort_cdp_responses(html_paths)
    for version in ordered_responses:
        html = ordered_responses[version]
        cdp_parser.parse_cdp_response(version, html, excel)
        print(name, version, "parsing completed")


company = "Naturgy Energy Group"
paths = ['file:///home/ivan/Documents/Zotero/storage/TEYH7E3R/pages.html',
         'file:///home/ivan/Documents/Zotero/storage/LJNED6WJ/pages.html',
         'file:///home/ivan/Documents/Zotero/storage/443EFXLY/responses.html',
         'file:///home/ivan/Documents/Zotero/storage/NPB36AS5/responses.html',
         'file:///home/ivan/Documents/Zotero/storage/3VJ4KSB6/responses.html']
parse_company(company, paths)
